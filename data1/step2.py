import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Load the CSV file
file_path = "stock_data.csv"  # Change if necessary
df = pd.read_csv(file_path)

# Convert Market Cap to numeric values for sorting
def convert_market_cap(value):
    if isinstance(value, str):
        value = value.replace("B", "e9").replace("M", "e6").replace("T", "e12")  # Convert Billion/Million/Trillion
        try:
            return float(eval(value))  # Convert formatted string to number
        except:
            return 0
    return value

df["Market Cap"] = df["Market Cap"].apply(convert_market_cap)

# Extract the first four rows that should not be sorted
fixed_rows = df.iloc[:4]  # Keep QQQ, SPY, IWM, AAPL at the top
sorted_rows = df.iloc[4:].sort_values(by="Market Cap", ascending=False)  # Sort remaining rows by Market Cap

# Combine fixed rows and sorted rows
df_sorted = pd.concat([fixed_rows, sorted_rows])

# Convert RSI, YTD%, and SMA values to numeric
df_sorted["RSI"] = pd.to_numeric(df_sorted["RSI"], errors="coerce")
df_sorted["YTD %"] = pd.to_numeric(df_sorted["YTD %"], errors="coerce")
df_sorted["20 SMA"] = pd.to_numeric(df_sorted["20 SMA"], errors="coerce")
df_sorted["50 SMA"] = pd.to_numeric(df_sorted["50 SMA"], errors="coerce")
df_sorted["200 SMA"] = pd.to_numeric(df_sorted["200 SMA"], errors="coerce")

# Save sorted data to Excel
excel_file = "final_stock_data.xlsx"
df_sorted.to_excel(excel_file, index=False)

# Open the Excel file for formatting
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Define fill colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Strong Green
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Strong Red
light_green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Light Green
light_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light Red

# Find column indexes for RSI, YTD%, and EMA levels
rsi_col = None
ytd_col = None
ema_20_col = None
ema_50_col = None
ema_200_col = None

for col_index, col_name in enumerate(df_sorted.columns, start=1):
    if col_name == "RSI":
        rsi_col = col_index
    elif col_name == "YTD %":
        ytd_col = col_index
    elif col_name == "20 SMA":
        ema_20_col = col_index
    elif col_name == "50 SMA":
        ema_50_col = col_index
    elif col_name == "200 SMA":
        ema_200_col = col_index

# Apply conditional formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
    if rsi_col:
        rsi_value = row[rsi_col - 1].value  # Get RSI value
        if isinstance(rsi_value, (int, float)):
            if rsi_value > 70:
                row[rsi_col - 1].fill = green_fill
            elif 65 <= rsi_value <= 70:
                row[rsi_col - 1].fill = light_green_fill
            elif 30 <= rsi_value <= 35:
                row[rsi_col - 1].fill = light_red_fill
            elif rsi_value < 30:
                row[rsi_col - 1].fill = red_fill

    if ytd_col:
        ytd_value = row[ytd_col - 1].value  # Get YTD % value
        if isinstance(ytd_value, (int, float)):
            if ytd_value > 15:
                row[ytd_col - 1].fill = green_fill
            elif ytd_value < -5:
                row[ytd_col - 1].fill = red_fill

    # Highlight EMA levels based on percentage difference
    for ema_col in [ema_20_col, ema_50_col, ema_200_col]:
        if ema_col:
            ema_value = row[ema_col - 1].value  # Get EMA percentage difference
            if isinstance(ema_value, (int, float)):
                if ema_value > 5:
                    row[ema_col - 1].fill = green_fill  # Strong Uptrend
                elif 0 < ema_value <= 5:
                    row[ema_col - 1].fill = light_green_fill  # Weak Uptrend
                elif -5 <= ema_value < 0:
                    row[ema_col - 1].fill = light_red_fill  # Weak Downtrend
                elif ema_value < -5:
                    row[ema_col - 1].fill = red_fill  # Strong Downtrend

# Adjust column widths dynamically
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the formatted Excel file
wb.save(excel_file)

print(f"Sorted and formatted data saved to {excel_file}")
